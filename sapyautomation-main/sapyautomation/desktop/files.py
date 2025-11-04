import shutil
from datetime import datetime
from pathlib import Path
from configparser import ConfigParser
import openpyxl
try:
    from PyPDF2 import PdfFileWriter, \
        PdfFileReader  # pylint: disable=import-error
except ModuleNotFoundError:
    pass

from sapyautomation.core.utils.strings import TIMESTAMP_FORMAT
import os
import pandas as pd
import tempfile

DIR_SEPARATORS = ('\\', '/')


def merge_pdfs(input_paths: list, output_path: str):
    pdf_writer = PdfFileWriter()
    for path in input_paths:
        pdf_reader = PdfFileReader(path)
        for page in range(pdf_reader.getNumPages()):
            pdf_writer.addPage(pdf_reader.getPage(page))

    with open(output_path, 'wb') as fh:
        pdf_writer.write(fh)


def is_relative_path(path_string: str) -> bool:
    """ Validates if path string is a relative or absolute path

    This method detects as absolute path strings starting with separators
    """
    return path_string[1] != ':' and path_string[0] not in DIR_SEPARATORS or \
        path_string[0] == '.' and path_string[1] in DIR_SEPARATORS


def test_path_format(path_string: str) -> str:
    """ Converts path string to a readable path for the framework

    Args:
        path_string(str): relative-path string to be converted

    """
    path_string = path_string.replace(DIR_SEPARATORS[0], DIR_SEPARATORS[1])

    if path_string[0] != '.':
        for i, char in enumerate(path_string):
            if char not in DIR_SEPARATORS:
                path_string = path_string[i:]
                break
            i = i+1

        path_string = '.%s%s' % (DIR_SEPARATORS[1], path_string)

    return path_string


def module_to_path(module_string: str) -> str:
    """ Converts module.string.like.this to a readable path for the framework

    Args:
        module_string(str): module string to be converted.

    """
    path_string = test_path_format('/'.join(module_string.split('.')))

    return path_string


def path_to_module(path_string: str) -> str:
    """ Converts path string to module.string.like.this

    path_string(str): relative-path to be converted.

    """
    path_string = test_path_format(path_string)[2:]
    parts = path_string.split(DIR_SEPARATORS[1])
    if '.' in parts[-1]:
        if '.py' in parts[-1]:
            parts[-1] = parts[-1][:-3]
        else:
            parts = parts[:-2]

    return '.'.join(parts)


def generate_unique_filename(file_path: str, add_timestamp: bool = False):
    """ Generates unique filename adding index and optional timestamp
    Args:
        file_path(str): full path to the file
        add_timestamp(bool): flag to indicate if timestamp is wanted
    """
    path = Path(file_path).parent
    base_filename = Path(file_path).parts[-1]
    filename = base_filename
    index = 0
    i_size = 1

    if base_filename[0].isdigit() and base_filename[1].isdigit():
        base_filename = base_filename[2:]
    if add_timestamp:
        timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
        filename = '%s_%s' % (timestamp, filename)

    while path.joinpath(filename).exists():
        index += 1
        if index > 9:
            i_size = 2

        if i_size == 1:
            filename = '0%s%s' % (index, base_filename)
        else:
            filename = '%s%s' % (index, base_filename)

    return path.joinpath(filename)


def generate_unique_dirname(dir_path: str, add_timestamp: bool = False):
    """ Generates unique dirname adding index and optional timestamp
    Args:
        file_path(str): full path to the file
        add_timestamp(bool): flag to indicate if timestamp is wanted
    """
    return generate_unique_filename(dir_path, add_timestamp)


class ConfigFile:
    """ Easy configuration files management

    If config file doesn't exists a copy of template file
    will be used when defined.

    Args:
        path(str): absolute path to the config file
        template_path(str):absolute path to template file for configuration.
    """
    def __init__(self, path: str, template_path: str = None):
        file_path = Path(path)
        if not file_path.exists() and template_path is not None:
            file_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy(template_path, path)

        self.__parser = ConfigParser()
        self.__parser.read(path)
        self.__path = path

    def get_bool(self, section: str, key: str):
        """ Gets key as boolean

        Args:
            key(str): key name to be retrieved.
            section(str): section's name in which key will be located.
        """
        value = self.__parser.get(section, key)
        if value.lower() == "false":
            value = False

        elif value.lower() == "true":
            value = True

        return value

    def get_int(self, section: str, key: str):
        """ Gets key as integer

        Args:
            key(str): key name to be retrieved.
            section(str): section's name in which key will be located.
        """
        value = self.__parser.get(section, key)
        try:
            return float(value)

        except ValueError:
            return None

    def get_float(self, section: str, key: str):
        """ Gets key as float

        Args:
            key(str): key name to be retrieved.
            section(str): section's name in which key will be located.
        """
        value = self.__parser.get(section, key)
        try:
            return int(value)

        except ValueError:
            return None

    def get_str(self, section: str, key: str):
        """ Gets key as string

        Args:
            key(str): key name to be retrieved.
            section(str): section's name in which key will be located.
        """
        value = self.__parser.get(section, key)
        if value.lower() == "none":
            return None

        return str(value)


class SpreadSheetData:
    """ Retrieves data from spreadsheet files
    suported formats(xls, xlsx)

    Args:
        path(str): path to spreadsheet file
    """

    def __init__(self, path: str):
        self.__path = path
        if path.endswith('.xls'):
            # save the xls workbook
            with tempfile.TemporaryDirectory() as temp_dir:
                # Concatenate the temporary directory path and the filename
                temp_path = os.path.join(temp_dir, f"{os.path.splitext(os.path.basename(path))[0]}.xlsx")
                #Save the XLSX file to an XLS file
                # Create a Pandas Excel writer using XlsxWriter as the engine.
                with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                    xls = pd.read_excel(path, sheet_name=None)
                    for sheet_name, df in xls.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                self._book = openpyxl.load_workbook(temp_path)

        else:
            self._book = openpyxl.load_workbook(path)

    def get_row_by_value(self, col_index: int, value: str,
                         sheet_index: int = 0):
        for i in range(self.count_rows(sheet_index)):
            cell_value = self.get_cell_data(col_index, i, sheet_index)

            if cell_value == value:
                return self.get_row_data(i, sheet_index)

        return None

    def count_rows(self, sheet_index: int = 0):
        """ Counts rows in specific sheet of the file

        Args:
            sheet_index(int): index of the sheet
        """
        sheet = self._book.worksheets[sheet_index]

        return sheet.max_row

    def count_cols(self, sheet_index: int = 0):
        """ Counts columns in specific sheet of the file

        Args:
            sheet_index(int): index of the sheet
        """
        sheet = self._book.worksheets[sheet_index]

        return sheet.max_column

    def get_sheets(self):
        """ Returns all sheets data from actual book
        """
        return self._book.sheetnames

    def get_column_data(self, column_index: int, sheet_index: int = 0) -> list:
        """ Gets all cells in specific column

        Args:
            column_index(int): index of the column
            sheet_index(int): index of the sheet

        Returns:
            cells list
        """
        cells = []

        for y in range(self.count_rows(sheet_index)):
            cells.append(self.get_cell_data(column_index, y, sheet_index))

        return cells

    def get_row_data(self, row_index: int, sheet_index: int = 0) -> list:
        """ Gets all cells in specific row

        Args:
            column_index(int): index of the row
            sheet_index(int): index of the sheet

        Returns:
            cells list
        """
        sheet = self._book.worksheets[sheet_index]
        cells = []

        for x in range(1, self.count_cols(sheet_index) + 1):
            cells.append(sheet.cell(row=row_index + 1, column=x).value)

        return cells

    def get_matrix(self, sheet_index: int = 0) -> list:
        """ Gets all cells from sheet

        Args:
            sheet_index(int): index of the sheet

        Returns:
            cells list
        """
        matrix = self.get_custom_matrix((0, 0),
                                        (self.count_cols(sheet_index),
                                         self.count_rows(sheet_index)),
                                        sheet_index)

        return matrix

    def get_custom_matrix(self, top_left: tuple, bottom_right: tuple,
                          sheet_index: int = 0) -> list:
        """ Gets cells from splitted matrix

        Args:
            top_left(tuple): top left corner of the custom matrix(x, y)
            bottom_right(tuple): bottom right corner of the custom matrix(x, y)
            sheet_index(int): index of the sheet

        Returns:
            cells list
        """
        cells = []
        for col_index in range(top_left[0], bottom_right[0]):
            col = self.get_column_data(col_index, sheet_index)
            cells.append(col[top_left[1]: bottom_right[1]])

        return cells

    def get_cell_data(self, x: int, y: int, sheet_index: int = 0):
        """ Gets data from specific cell

        Args:
            x(int): column index
            y(int): row index
            sheet_index(int): index of the sheet
        """
        row = self.get_row_data(y, sheet_index)

        return row[x]

    def append_row(self, cells_data: list, sheet_index: int = 0,
                   offset: int = 0):
        """ Adds at the end of a specific sheet a row of data
        Args:
            cells_data(list): Cell-data array to be appended
            sheet_index(int): index of the sheet
        """
        book = openpyxl.load_workbook(self.__path)
        sheet = book.worksheets[sheet_index]
        new_row = self.count_rows(sheet_index) + offset

        for cell, data in enumerate(cells_data):
            sheet.cell(row=new_row+1, column=cell+1).value = data

        book.save(self.__path)

    def append_column(self, cells_data: list, sheet_index: int = 0):
        """ Adds at the end of a specific sheet a column of data
        Args:
            cells_data(list): Cell-data array to be appended
            sheet_index(int): index of the sheet
        """
        book = openpyxl.load_workbook(self.__path)
        sheet = book.worksheets[sheet_index]
        new_col = self.count_cols(sheet_index)

        for cell, data in enumerate(cells_data):
            sheet.cell(row=cell+1, column=new_col+1).value = data

        book.save(self.__path)
